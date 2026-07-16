import logging
import io
import time
from typing import Tuple, List, Dict, Any, AsyncIterator
from openpyxl import load_workbook

from core.exceptions import FileProcessingException
from .base_extractor import ITextExtractor

logger = logging.getLogger(__name__)

class ExcelTextExtractor(ITextExtractor):
    """Extract text from Excel spreadsheet."""

    async def extract_legacy(self, content: bytes) -> Tuple[List[str], float, Dict[str, Any]]:
        if not content:
            raise FileProcessingException("excel", "Empty XLSX provided")
        
        try:
            start_time = time.time()
            wb = load_workbook(io.BytesIO(content))
            
            all_text = []
            # for sheet in wb.sheetnames[:5]:  # First 5 sheets
            for sheet in wb.sheetnames:  # All sheets
                ws = wb[sheet]
                sheet_texts = []
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and str(cell).strip():
                            sheet_texts.append(str(cell).strip())
                if sheet_texts:
                    all_text.append(f"Sheet {sheet}: " + " ".join(sheet_texts))
            
            confidence = 1.0 if all_text else 0.0
            processing_time = time.time() - start_time
            
            log_entry = {
                "source": "excel_native",
                "success": True,
                # "sheets_processed": min(5, len(wb.sheetnames)),
                "sheets_processed": len(wb.sheetnames),
                "text_length": sum(len(t) for t in all_text),
                "confidence": confidence,
                "processing_time": processing_time
            }
            
            return all_text, confidence, log_entry
        except Exception as e:
            logger.error(f"Excel extraction failed: {e}")
            raise FileProcessingException("excel", f"Extraction failed: {e}")

    async def extract_stream(self, content: bytes) -> AsyncIterator[Tuple[str, float, Dict[str, Any]]]:
        if not content:
            raise FileProcessingException("excel", "Empty XLSX provided")
            
        try:
            wb = load_workbook(io.BytesIO(content))
            total_sheets = len(wb.sheetnames)
            for sheet_idx, sheet in enumerate(wb.sheetnames):
                ws = wb[sheet]
                sheet_texts = []
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and str(cell).strip():
                            sheet_texts.append(str(cell).strip())
                text = " ".join(sheet_texts).strip()
                if text:
                    logger.info(f"Extracted native text from Excel sheet '{sheet}': {text[:100]}...")
                    yield text, 1.0, {
                        "source": "excel_native_sheet",
                        "success": True,
                        "sheet_name": sheet,
                        "sheet_index": sheet_idx + 1,
                        "total_sheets": total_sheets,
                        "text_length": len(text),
                    }
        except Exception as e:
            logger.error(f"Excel extraction failed: {e}")
            raise FileProcessingException("excel", f"Extraction failed: {e}")
